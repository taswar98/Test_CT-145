import requests
import re
import pandas as pd
import os
import xlsxwriter
import glob
import scraper
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from urllib.request import urlopen
from bs4 import BeautifulSoup
from datetime import date
from pathlib import Path

############################################################
# This is write header to excel file
headers = ['Bank', 'Account Name', 'Account Type', 'Monthly Fee', 'Special Offer', 'Expiry Date', 'Account Perks',
           'Webiste']
book = Workbook()
sheet = book.active

for index, header in enumerate(headers):
    sheet.cell(row=1, column=index + 1).font = Font(bold=True)
    sheet.cell(row=1, column=index + 1).value = header

# 1. Get dictionary return value from scraper.py for sepecial offer.
# print(json.dumps(scraper.get_special_offer_accounts(), indent=1))
special_offer_dict = scraper.get_special_offer_accounts()


# print(dic)

def getList(dict):
    return special_offer_dict.keys()


# print(getList(dict))

# print("dict "+ len(dict.keys()))
# Get bank name
# print(dict[0]['institution_name'])

rowNum = 2
for x in getList(special_offer_dict):
    # Get Bank name
    bankName = special_offer_dict[x]['institution_name']

    for y in range(len(special_offer_dict[x]['accounts'])):
        sheet.cell(row=rowNum, column=1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=rowNum, column=1).value = bankName

        sheet.cell(row=rowNum, column=2).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=rowNum, column=2).value = special_offer_dict[x]['accounts'][y]['account_name'][0]

        sheet.cell(row=rowNum, column=3).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=rowNum, column=3).value = special_offer_dict[x]['accounts'][y]['account_category']

        sheet.cell(row=rowNum, column=4).alignment = Alignment(wrapText=True)
        # sheet.cell(row=rowNum, column=4).alignment = Alignment(horizontal='left', vertical='center')
        writtenPara = ""
        if not special_offer_dict[x]['accounts'][y]['fee']:
            writtenPara = "$0"
        else:
            for index, fee in enumerate(special_offer_dict[x]['accounts'][y]['fee']):
                writtenPara += str(index + 1) + ". " + fee + "\n"
        sheet.cell(row=rowNum, column=4).value = writtenPara

        sheet.cell(row=rowNum, column=5).alignment = Alignment(wrapText=True)
        writtenPara = ""
        if not special_offer_dict[x]['accounts'][y]['special_offer']:
            writtenPara = "No Data!"
        else:
            for index, special_offer in enumerate(special_offer_dict[x]['accounts'][y]['special_offer']):
                writtenPara += str(index + 1) + ". " + special_offer + "\n"
        sheet.cell(row=rowNum, column=5).value = writtenPara
        # sheet.cell(row=rowNum, column=4).value=special_offer_dict[x]['accounts'][y]['fee']

        writtenPara = ""
        for index, detail in enumerate(special_offer_dict[x]['accounts'][y]['details']):
            writtenPara += str(index + 1) + ". " + detail + "\n"
        sheet.cell(row=rowNum, column=7).value = writtenPara

        sheet.cell(row=rowNum, column=8).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=rowNum, column=8).value = r'=HYPERLINK("http://www.example.com","' + bankName + '")'
        print(bankName)

        rowNum += 1

todayDate = str(date.today())

book.save("specialOffer" + todayDate + ".xlsx")