import requests
import re
import pandas as pd
import os
import xlsxwriter
import glob
import scraper
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from urllib.request import urlopen
from bs4 import BeautifulSoup
from datetime import date
import string
from pathlib import Path

headers = ['Bank', 'Account Name', 'Account Type', 'Monthly Fee', 'Special Offer', 'Expiry Date', 'Account Perks',
           'Website']
book = Workbook()
sheet = book.active

for index, header in enumerate(headers):
    sheet.cell(row=1, column=index + 1).font = Font(bold=True)
    sheet.cell(row=1, column=index + 1).value = header

special_offer_dict = scraper.get_special_offer_accounts()


def getList(dict):
    return special_offer_dict.keys()


rowNum = 2
for x in getList(special_offer_dict):
    # Get Bank name
    bankName = special_offer_dict[x]['institution_name']

    for y in range(len(special_offer_dict[x]['accounts'])):
        sheet.cell(row=rowNum, column=1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=rowNum, column=1).value = bankName

        sheet.cell(row=rowNum, column=2).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=rowNum, column=2).value = special_offer_dict[x]['accounts'][y]['account_name'][0]

        sheet.cell(row=rowNum, column=3).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=rowNum, column=3).value = special_offer_dict[x]['accounts'][y]['account_category']

        sheet.cell(row=rowNum, column=4).alignment = Alignment(wrapText=True)
        # sheet.cell(row=rowNum, column=4).alignment = Alignment(horizontal='left', vertical='center')
        writtenPara = ""
        if not special_offer_dict[x]['accounts'][y]['fee']:
            writtenPara = "$0"
        else:
            for index, fee in enumerate(special_offer_dict[x]['accounts'][y]['fee']):
                writtenPara += str(index + 1) + ". " + fee + "\n"
        sheet.cell(row=rowNum, column=4).value = writtenPara

        sheet.cell(row=rowNum, column=5).alignment = Alignment(wrapText=True)
        writtenPara = ""
        if not special_offer_dict[x]['accounts'][y]['special_offer']:
            writtenPara = "No Data!"
        else:
            for index, special_offer in enumerate(special_offer_dict[x]['accounts'][y]['special_offer']):
                if bankName == 'Scotiabank':
                    if index == 6:
                        break
                det1 = str(special_offer)
                cleaned = det1.replace('legal bug', '').rstrip(string.digits)
                # print("$$$"+cleaned)
                writtenPara += str(index + 1) + ". " + cleaned + "\n"
        sheet.cell(row=rowNum, column=5).value = writtenPara
        # sheet.cell(row=rowNum, column=4).value=special_offer_dict[x]['accounts'][y]['fee']

        writtenPara = ""
        for index, detail in enumerate(special_offer_dict[x]['accounts'][y]['details']):
            det = str(detail)
            cleaned = det.replace('legal bug', '').rstrip(string.digits)
            writtenPara += str(index + 1) + ". " + cleaned + "\n"
        sheet.cell(row=rowNum, column=7).value = writtenPara

        sheet.cell(row=rowNum, column=8).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=rowNum, column=8).value = r'=HYPERLINK("http://www.example.com","' + bankName + '")'
        print(bankName)

        rowNum += 1

############################################################
# The list of all potential FI's websites that we might visit
pages = ["https://www.cibc.com/en/special-offers/fall-savings-promotion.html",
         "https://www.scotiabank.com/ca/en/personal/rates-prices/savings-account-rates.html",
         # "https://www.scotiabank.com/ca/en/personal/bank-accounts/savings-accounts/momentum-plus-savings-account.html",
         "https://www.tangerine.ca/en/landing-page/raptors"]

# Today date in order to generate Excel timestamp
todayDate = str(date.today())

book.save("specialOffer" + todayDate + ".xlsx")
####################################################

# def compare_changed_Special_Offer(previousFile,todayFile):


# Today date in order to generate Excel timestamp
todayDate = str(date.today())

# Get correct file
excelFile = glob.glob("specialOffer[0-9]*.xlsx")
excelFile = sorted(excelFile)
# path to files
currentDirectory = os.getcwd() + "\\"
path_OLD = Path(currentDirectory + excelFile[len(excelFile) - 2])
path_NEW = Path(currentDirectory + excelFile[len(excelFile) - 1])

# Read in the two excel files and fill NA
df_OLD = pd.read_excel(path_OLD, header=None, names=None).fillna(0)
df_NEW = pd.read_excel(path_NEW, header=None, names=None).fillna(0)

dfDiff = df_OLD.copy()
for row in range(dfDiff.shape[0]):
    for col in range(dfDiff.shape[1]):
        value_OLD = df_OLD.iloc[row, col]
        try:
            value_NEW = df_NEW.iloc[row, col]
            if value_OLD == value_NEW and value_NEW != 0:
                dfDiff.iloc[row, col] = df_NEW.iloc[row, col]
            elif value_OLD == value_NEW and value_NEW == 0:
                dfDiff.iloc[row, col] = ""
            elif (value_OLD != value_NEW and value_NEW == 0):
                dfDiff.iloc[row, col] = ('Expired: {}').format(value_OLD)
            else:
                dfDiff.iloc[row, col] = ('Update: {}').format(value_NEW)

        except:
            dfDiff.iloc[row, col] = ('{}-->{}').format(value_OLD, 'NaN')
writer = pd.ExcelWriter("specialOffer_compare" + todayDate + ".xlsx",
                        engine='xlsxwriter')  # pylint: disable=abstract-class-instantiated
dfDiff.to_excel(writer, sheet_name='DIFF', index=False, header=None)

workbook = writer.book
worksheet = writer.sheets['DIFF']

# define formats
highlight_fmt_red = workbook.add_format({'font_color': '#000000', 'bg_color': '#FF0000'})
highlight_fmt_yellow = workbook.add_format({'font_color': '#000000', 'bg_color': '#FFFF00'})

## highlight Update cells
worksheet.conditional_format('A1:ZZ1000', {'type': 'text',
                                           'criteria': 'containing',
                                           'value': 'Update',
                                           'format': highlight_fmt_yellow})
## highlight Expired cells
worksheet.conditional_format('A1:ZZ1000', {'type': 'text',
                                           'criteria': 'containing',
                                           'value': 'Expired',
                                           'format': highlight_fmt_red})
# save
writer.save()