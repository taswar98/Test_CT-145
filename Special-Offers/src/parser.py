import re
import xlsxwriter
from yaml_utils import YAMLUtils
from tqdm import tqdm
import openpyxl
from openpyxl import Workbook

banks = {0: {'institution_name': 'RBC', 'accounts': [{'account_category': 'Saving Accounts', 'account_name': ['RBC High Interest eSavings'], 'fee': ['$10'], 'details': ['High interst on every dollarnr banking account to your savings'], 'special_offer': ['Interest Rate: 0.050%']}]}}

def createcolumns():

    headers = ['Bank', 'Account', 'Details', 'Special Offer']
    book = Workbook()
    sheet = book.active

    for index, header in enumerate(headers):
        sheet.cell(row=1, column=index + 1).value = header

    book.save(filename='myoutput.xlsx')


def main():
    createcolumns()

    book = Workbook()
    sheet = book.active
    
    data = list(dictiomnaruy)
    for key,value in enumerate(banks):
        sheet.cell()





main()
